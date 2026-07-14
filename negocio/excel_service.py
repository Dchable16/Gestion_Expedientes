# -*- coding: utf-8 -*-
"""
Created on Tue Jul 22 11:38:36 2025

@author: dchable
"""

# negocio/excel_service.py

from openpyxl.cell.cell import MergedCell
import pandas as pd
import logging
import os

from typing import List, Dict, Any, Tuple
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter 
from openpyxl import load_workbook
from datetime import datetime
from copy import copy

from utils.config_manager import get_logo_semarnat_path, get_logo_conagua_path


class ExcelService:
    
    def create_report(self, data: List[Dict[str, Any]], filepath: str) -> Tuple[bool, str]:
        """
        Crea un reporte en formato Excel, asegurando que TODAS las fechas (incluyendo Gestión)
        tengan el formato día-mes-año.
        """
        import time
        time.sleep(5)
        
        if not data:
            return False, "No hay datos para exportar."
        if not filepath:
            return False, "No se especificó una ruta para guardar el archivo."

        try:
            df = pd.DataFrame(data)
            numeric_cols = ['ID', 'ID Expediente', 'ID Respuesta', 'Páginas', 'Apertura', 'Cierre']
            date_cols = [
                'Fecha', 'Vencimiento', 'Fecha Respuesta', 'Fecha Ingreso', 
                'Fecha Recepción', 'Fecha Documento', 'Fecha Límite' 
            ]

            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            for col in date_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d-%m-%Y').fillna('')
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Reporte', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Reporte']
                alignment_style = Alignment(wrap_text=True, vertical='top')
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = alignment_style
                for i, column_cells in enumerate(worksheet.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(i)
                    
                    for cell in column_cells:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
            
            return True, f"Reporte guardado exitosamente en: {os.path.abspath(filepath)}"

        except Exception as e:
            logging.error("Error al crear reporte Excel: %s", e, exc_info=True)
            return False, f"Ocurrió un error al generar el reporte: {e}"
    
    def create_inventory_from_template(self, data: List[Dict], template_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Crea un reporte de inventario usando una plantilla.
        Este método orquesta la clonación de hojas, inserción dinámica de filas 
        para no borrar los pies de página y la reubicación exacta de firmas.
        """
        import io
        from copy import deepcopy
        from openpyxl.drawing.image import Image as OpenpyxlImage
        
        if not data:
            return False, "No hay datos para generar el inventario."

        try:
            df = pd.DataFrame(data)
            book = load_workbook(template_path)
            template_sheet_name = book.sheetnames[1]
            template_sheet = book[template_sheet_name]

            self._poblar_hoja_resumen(book.worksheets[0], df)

            grouped = df.groupby('serie_documental')
            for serie_name, group_df in grouped:
                if not serie_name or serie_name == "Conocimiento":
                    continue

                new_sheet = book.copy_worksheet(template_sheet)
                new_sheet.title = str(serie_name)[:31]
                cantidad_registros = len(group_df)
                filas_extra = cantidad_registros - 1 if cantidad_registros > 0 else 0
                if filas_extra > 0:
                    new_sheet.insert_rows(14, amount=filas_extra)
                for imagen in template_sheet._images:
                    try:
                        nuevo_anchor = deepcopy(imagen.anchor)
                        tipo_ancla = type(nuevo_anchor).__name__
                        if tipo_ancla in ['TwoCellAnchor', 'OneCellAnchor'] and nuevo_anchor._from.row >= 12:
                            img_bytes = io.BytesIO()
                            imagen.ref.seek(0)
                            img_bytes.write(imagen.ref.read())
                            img_bytes.seek(0)
                            
                            nueva_imagen = OpenpyxlImage(img_bytes)
                            nuevo_anchor._from.row += filas_extra
                            if tipo_ancla == 'TwoCellAnchor':
                                if hasattr(nuevo_anchor, '_to'):
                                    nuevo_anchor._to.row += filas_extra
                                elif hasattr(nuevo_anchor, 'to'):
                                    nuevo_anchor.to.row += filas_extra
                            nueva_imagen.anchor = nuevo_anchor
                            new_sheet.add_image(nueva_imagen)
                    except Exception as e:
                        logging.warning(f"Error al clonar y reubicar imagen de firma: {e}")
                self._insertar_logos_en_hoja(new_sheet)
                self._escribir_cabeceras_serie(new_sheet, group_df, serie_name)
                self._escribir_datos_expedientes(new_sheet, group_df)

            del book[template_sheet_name]
            book.save(output_path)
            
            return True, f"Inventario generado exitosamente en: {os.path.abspath(output_path)}"

        except Exception as e:
            logging.error("Error al crear inventario desde plantilla: %s", e, exc_info=True)
            return False, f"Ocurrió un error al generar el inventario: {e}"

    def _poblar_hoja_resumen(self, sheet, df: pd.DataFrame):
        """Escribe los datos del resumen general en la primera hoja del Excel."""
        summary_columns = [
            'id', 'tipo_documento', 'categoria_documental', 'folio', 'fecha', 
            'asunto', 'serie_documental', 'carpeta', 'paginas', 'clasificacion', 
            'apertura', 'cierre', 'vencimiento'
        ]
        summary_columns_existentes = [col for col in summary_columns if col in df.columns]
        summary_df = df[summary_columns_existentes]
        
        for c_idx, header in enumerate(summary_df.columns, 1):
            sheet.cell(row=1, column=c_idx, value=header)
        for r_idx, row in enumerate(summary_df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _insertar_logos_en_hoja(self, sheet):
        """Inserta los logos de SEMARNAT y CONAGUA en la hoja especificada."""
        try:
            logo1_path = get_logo_semarnat_path()
            logo2_path = get_logo_conagua_path()
            
            if os.path.exists(logo1_path):
                img1 = Image(logo1_path)
                sheet.add_image(img1, 'A2')
            else:
                logging.warning(f"No se encontró el logo de SEMARNAT en: {logo1_path}")

            if os.path.exists(logo2_path):
                img2 = Image(logo2_path)
                sheet.add_image(img2, 'P2')
            else:
                logging.warning(f"No se encontró el logo de CONAGUA en: {logo2_path}")
        except Exception:
            logging.error("Ocurrió un error al intentar insertar los logos.", exc_info=True)

    def _escribir_cabeceras_serie(self, sheet, group_df: pd.DataFrame, serie_name: str):
        """Escribe la información de la sección y la serie en las celdas C9 y C10."""
        if group_df.empty:
            return
            
        seccion_valor = group_df['seccion'].iloc[0]
        celda_seccion = sheet['C9']
        celda_seccion.value = seccion_valor
        celda_seccion.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        nombre_serie = group_df['nombre_serie'].iloc[0]
        texto_celda_c10 = f"{serie_name} {nombre_serie}"
        celda_serie = sheet['C10']
        celda_serie.value = texto_celda_c10
        celda_serie.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
    def _escribir_datos_expedientes(self, sheet, group_df: pd.DataFrame):
        """
        Escribe los datos en la plantilla respetando las 24 columnas.
        Aplica bordes, ajuste de texto y alineación profesional.
        """
        from openpyxl.styles import Border, Side, Alignment
        
        if group_df.empty:
            return
        
        borde_fino = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        alineacion_centro = Alignment(wrap_text=True, vertical='center', horizontal='center')
        alineacion_izq = Alignment(wrap_text=True, vertical='center', horizontal='left')
        
        def extraer_numero_exp(clasificacion):
            if not isinstance(clasificacion, str): return 0
            partes = clasificacion.split('/')
            return int(partes[2]) if len(partes) >= 4 and partes[2].isdigit() else 0

        group_df['num_exp_temp'] = group_df['clasificacion'].apply(extraer_numero_exp)
        group_df.sort_values(by=['apertura', 'num_exp_temp'], ascending=[False, True], inplace=True)
        def marca(valor):
            return "X" if str(valor).strip().upper() in ["SI", "X", "1", "TRUE"] else ""
        for r_idx, (_, exp) in enumerate(group_df.iterrows(), start=13):
            datos_fila = [
                r_idx - 12,                                   # 1. Consecutivo
                exp.get('lote_origen', ''),                   # 2. Número de caja
                extraer_numero_exp(exp.get('clasificacion')), # 3. Número de expediente
                exp.get('clasificacion', ''),                 # 4. Clasificación Archivística
                exp.get('asunto', ''),                        # 5. Descripción o asunto
                exp.get('apertura', ''),                      # 6. Año apertura
                exp.get('cierre', ''),                        # 7. Año cierre
                exp.get('paginas', 0),                        # 8. Número de folios
                marca(exp.get('administrativo')),             # 9. Valor: Admin
                marca(exp.get('legal')),                      # 10. Valor: Legal
                marca(exp.get('fiscal')),                     # 11. Valor: Fiscal
                exp.get('tramite', 0),                        # 12. Vigencia: Trámite
                exp.get('concentracion', 0),                  # 13. Vigencia: Concentración
                exp.get('total', 0),                          # 14. Vigencia: Total
                marca(exp.get('reservada')),                  # 15. Acceso: Reservada
                marca(exp.get('confidencial')),               # 16. Acceso: Confidencial
                marca(exp.get('publica')),                    # 17. Acceso: Pública
                exp.get('ubicacion_area', ''),                # 18. Ubicación: Área
                exp.get('ubicacion_pasillo', ''),             # 19. Ubicación: Pasillo
                exp.get('ubicacion_anaquel', ''),             # 20. Ubicación: Anaquel
                exp.get('ubicacion_charola', ''),             # 21. Ubicación: Charola
                marca(exp.get('original')),                   # 22. Tradición: Original
                marca(exp.get('copia')),                      # 23. Tradición: Copia
                exp.get('')                       # 24. No. de Tomo
            ]

            for c_idx, valor in enumerate(datos_fila, start=1):
                celda = sheet.cell(row=r_idx, column=c_idx, value=valor)
                celda.border = borde_fino
                if c_idx in [4, 5]:
                    celda.alignment = alineacion_izq
                else:
                    celda.alignment = alineacion_centro
                
    def _escribir_celda_segura(self, ws, coordenada, valor, ajustar_texto=False):
        """
        Escribe en una celda manejando uniones (merged cells) y permite activar
        el ajuste de texto para campos multilínea.
        """
        celda = ws[coordenada]
        target_cell = celda
        if isinstance(celda, MergedCell):
            for rango in ws.merged_cells.ranges:
                if coordenada in rango:
                    target_cell = ws.cell(row=rango.min_row, column=rango.min_col)
                    break
        target_cell.value = valor
        if ajustar_texto:
            target_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    def _formatear_fecha(self, fecha_str):
        """Convierte YYYY-MM-DD a DD-MM-YYYY."""
        if not fecha_str or not isinstance(fecha_str, str):
            return ""
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
            return fecha_obj.strftime('%d-%m-%Y')
        except ValueError:
            return fecha_str

    def generar_volante_cg(self, datos: Dict[str, Any], template_path: str, output_path: str) -> Tuple[bool, str]:
        """
        Rellena la plantilla de Control de Gestión, limpia las casillas de selección
        y marca las opciones correctas de forma segura (celdas combinadas).
        """
        try:
            if not os.path.exists(template_path):
                return False, f"No se encontró la plantilla en: {template_path}"

            wb = load_workbook(template_path)
            ws = wb.active
            celdas_opciones = [
                'L8', 'L9',          # Prioridad (Urgente, Normal)
                'L15', 'L16', 'L17', 'L18', # Tipo Instrucción
                'L35', 'L36',        # Anexos (Si, No)
                'L38', 'L39'         # Requiere Respuesta (Si, No)
            ]
            for celda in celdas_opciones:
                self._escribir_celda_segura(ws, celda, "") # Escribir vacío

            self._escribir_celda_segura(ws, 'J2', datos.get('folio', ''))
            self._escribir_celda_segura(ws, 'J5', self._formatear_fecha(datos.get('fecha', '')))
            
            self._escribir_celda_segura(ws, 'B7', datos.get('turnado_a', ''))
            self._escribir_celda_segura(ws, 'J10', self._formatear_fecha(datos.get('fecha_limite', '')))
            
            self._escribir_celda_segura(ws, 'B13', datos.get('remitente', ''))
            self._escribir_celda_segura(ws, 'B16', datos.get('area', ''))
            
            self._escribir_celda_segura(ws, 'B20', datos.get('referencia', ''))
            self._escribir_celda_segura(ws, 'G21', datos.get('detalle_instruccion', '')) 
            
            self._escribir_celda_segura(ws, 'B23', self._formatear_fecha(datos.get('fecha_documento', '')))
            self._escribir_celda_segura(ws, 'B26', datos.get('asunto', ''))

            self._escribir_celda_segura(ws, 'G28', datos.get('observaciones', ''))
            self._escribir_celda_segura(ws, 'B36', datos.get('recibio', ''))
            self._escribir_celda_segura(ws, 'C48', datos.get('ccp', ''))
            
            self._escribir_celda_segura(ws, 'B7', datos.get('turnado_a', ''), ajustar_texto=True) 
            self._escribir_celda_segura(ws, 'B13', datos.get('remitente', ''), ajustar_texto=True) 
            self._escribir_celda_segura(ws, 'G21', datos.get('detalle_instruccion', ''), ajustar_texto=True) 
            self._escribir_celda_segura(ws, 'B26', datos.get('asunto', ''), ajustar_texto=True)            
            self._escribir_celda_segura(ws, 'G28', datos.get('observaciones', ''), ajustar_texto=True)
            self._escribir_celda_segura(ws, 'B36', datos.get('recibio', ''), ajustar_texto=True)
            self._escribir_celda_segura(ws, 'C48', datos.get('ccp', ''), ajustar_texto=True)

            # --- FASE 3: MARCADO DE "X" (LÓGICA CONDICIONAL) ---
            
            # 1. Prioridad
            prioridad = str(datos.get('prioridad', '')).upper()
            if prioridad == 'URGENTE':
                self._escribir_celda_segura(ws, 'L7', 'X')
            else:
                self._escribir_celda_segura(ws, 'L8', 'X') # Normal por defecto

            # 2. Tipo de Instrucción
            instruccion = str(datos.get('tipo_instruccion', '')).upper()
            opciones_instruccion = {
                'ATENCIÓN PROCEDENTE': 'L14',
                'ATENCIÓN COORDINADA': 'L15',
                'ATENCIÓN GRUPAL': 'L16',
                'PARA SU CONOCIMIENTO': 'L17', # Coincide con tu ComboBox
                'CONOCIMIENTO': 'L17'          # Por si acaso llega simplificado
            }
            celda_instruccion = opciones_instruccion.get(instruccion)
            if celda_instruccion:
                self._escribir_celda_segura(ws, celda_instruccion, 'X')

            # 3. Documentos Anexos (L35=SI, L36=NO)
            anexos = str(datos.get('documentos_anexos', '')).upper()
            if anexos == 'SI':
                self._escribir_celda_segura(ws, 'L35', 'X')
            else:
                self._escribir_celda_segura(ws, 'L36', 'X')

            # 4. Requiere Respuesta (L38=SI, L39=NO)
            req_respuesta = str(datos.get('requiere_respuesta', '')).upper()
            if req_respuesta == 'SI':
                self._escribir_celda_segura(ws, 'L38', 'X')
            else:
                self._escribir_celda_segura(ws, 'L39', 'X')

            # Guardar el archivo final
            wb.save(output_path)
            return True, f"Formato generado correctamente en: {output_path}"

        except Exception as e:
            logging.error(f"Error al generar volante CG: {e}", exc_info=True)
            return False, f"Error al generar el Excel: {str(e)}"
    
    def convertir_excel_a_pdf(self, excel_path):
        """
        Usa Excel en segundo plano para guardar una copia en PDF.
        Requiere tener MS Excel instalado y la librería pywin32.
        """
        try:
            import win32com.client
            import pythoncom
            
            # Necesario si se ejecuta en hilos secundarios
            pythoncom.CoInitialize() 
            
            # Abrir instancia de Excel invisible
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Abrir el libro
            abs_path = os.path.abspath(excel_path)
            wb = excel.Workbooks.Open(abs_path)
            
            # Definir ruta PDF (cambiamos extensión .xlsx por .pdf)
            pdf_path = os.path.splitext(abs_path)[0] + ".pdf"
            
            # 0 = xlTypePDF
            wb.ExportAsFixedFormat(0, pdf_path)
            
            return True, pdf_path
            
        except ImportError:
            return False, "La librería 'pywin32' no está instalada."
        except Exception as e:
            return False, f"Error al convertir a PDF: {str(e)}"
        finally:
            if wb:
                try:
                    wb.Close(False)
                except:
                    pass
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
            import pythoncom
            pythoncom.CoUninitialize()
    
    def generar_etiquetas_excel(self, datos_completos, template_path, output_path):
        """
        Copia una plantilla de Excel y reemplaza las etiquetas {{CLAVE}} con los datos del expediente.
        """
        import openpyxl
        import logging
        
        try:
            expediente = datos_completos.get("expediente", {})
            respuestas = datos_completos.get("respuestas", []) # <--- Extraemos las respuestas
            serie = datos_completos.get("info_serie", {}) or {}
            total_folios = 0
            
            try:
                if expediente.get("paginas"):
                    total_folios += int(expediente.get("paginas"))
            except ValueError:
                pass
                
            for resp in respuestas:
                try:
                    if resp.get("paginas"):
                        total_folios += int(resp.get("paginas"))
                except ValueError:
                    pass

            # --- PREPARACIÓN DE LOS DEMÁS DATOS ---
            def marca(valor):
                return "X" if str(valor).strip().upper() in ["SI", "X", "1", "TRUE"] else ""

            clasificacion = expediente.get("clasificacion", "")
            num_exp = ""
            if clasificacion:
                partes = clasificacion.split('/')
                if len(partes) >= 3:
                    num_exp = partes[-2]

            reemplazos = {
                "{{ASUNTO}}": expediente.get("asunto", ""),
                "{{CLASIFICACION}}": clasificacion,
                "{{FOLIOS}}": str(total_folios), # <--- AQUÍ INYECTAMOS LA SUMA TOTAL
                "{{SECCION}}": serie.get('seccion', ''),
                "{{SERIE}}": f"{serie.get('codigo_serie', '')} {serie.get('nombre_serie', '')}".strip(),
                "{{X_PUB}}": marca(serie.get('publica')),
                "{{X_RES}}": marca(serie.get('reservada')),
                "{{X_CON}}": marca(serie.get('confidencial')),
                "{{X_ORI}}": marca(serie.get('original')),
                "{{X_COP}}": marca(serie.get('copia')),
                "{{X_ADM}}": marca(serie.get('administrativo')),
                "{{X_LEG}}": marca(serie.get('legal')),
                "{{X_FIS}}": marca(serie.get('fiscal')),
                "{{T_TRA}}": str(serie.get('tramite', '')),
                "{{T_CON}}": str(serie.get('concentracion', '')),
                "{{T_TOT}}": str(serie.get('total', '')),
                "{{EXPEDI}}": num_exp,
                "{{APERTURA}}": str(expediente.get("apertura", "")),
                "{{CIERRE}}": str(expediente.get("cierre", ""))
            }

            wb = openpyxl.load_workbook(template_path)
            
            for hoja in wb.worksheets:
                for fila in hoja.iter_rows():
                    for celda in fila:
                        if celda.value and isinstance(celda.value, str):
                            # Si la celda tiene texto, buscamos si hay alguna etiqueta para reemplazar
                            for clave, valor_real in reemplazos.items():
                                if clave in celda.value:
                                    texto_reemplazo = str(valor_real) if valor_real not in [None, "None"] else ""
                                    celda.value = celda.value.replace(clave, texto_reemplazo)
            
            wb.save(output_path)
            return True, output_path
            
        except Exception as e:
            logging.error(f"Error al generar etiquetas en Excel: {e}", exc_info=True)
            return False, str(e)
    
    def generar_inventario_transferencia(self, expediente_service, ids_expedientes, template_path, output_path):
        """
        Agrupa los expedientes por serie, clona la plantilla para cada serie,
        nombra las pestañas con el código de la serie y escribe los datos.
        Incluye manejo seguro de imágenes, desplazamiento dinámico de pies de página 
        y formateo completo (bordes de 24 columnas).
        """
        import openpyxl
        from openpyxl.styles import Alignment, Border, Side
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from copy import deepcopy
        import io
        import logging
        import re
        
        try:
            expedientes_agrupados = {}
            
            for exp_id in ids_expedientes:
                datos_completos = expediente_service.obtener_vista_completa_expediente(exp_id)
                if not datos_completos: 
                    continue
                    
                serie = datos_completos.get("info_serie", {}) or {}
                codigo_serie = serie.get("codigo_serie", "Sin_Serie")
                
                codigo_seguro = re.sub(r'[\\/*?:\[\]]', '_', codigo_serie)[:31]
                
                if codigo_seguro not in expedientes_agrupados:
                    expedientes_agrupados[codigo_seguro] = []
                    
                expedientes_agrupados[codigo_seguro].append(datos_completos)

            if not expedientes_agrupados:
                return False, "No se encontraron datos válidos para los expedientes seleccionados."

            wb = openpyxl.load_workbook(template_path)
            hoja_plantilla = wb.active 
            
            fila_inicio = 13 
            
            def marca(valor):
                return "X" if str(valor).strip().upper() in ["SI", "X", "1", "TRUE"] else ""

            # Definición del estilo de borde
            borde_fino = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

            for codigo_serie, lista_datos in expedientes_agrupados.items():
                
                nueva_hoja = wb.copy_worksheet(hoja_plantilla)
                nueva_hoja.title = codigo_serie
                
                # --- 1. CÁLCULO DE DESPLAZAMIENTO DINÁMICO ---
                cantidad_registros = len(lista_datos)
                filas_extra = cantidad_registros - 1 if cantidad_registros > 0 else 0

                # --- 2. EMPUJAR CELDAS DE TEXTO (PIE DE PÁGINA) HACIA ABAJO ---
                if filas_extra > 0:
                    nueva_hoja.insert_rows(14, amount=filas_extra)
                
                # --- 3. COPIAR Y REUBICAR IMÁGENES/FIRMAS ---
                for imagen in hoja_plantilla._images:
                    try:
                        nuevo_anchor = deepcopy(imagen.anchor)
                        tipo_ancla = type(nuevo_anchor).__name__
                        
                        # Si la imagen está en la zona de registros o pie de página (>= fila 13)
                        if tipo_ancla in ['TwoCellAnchor', 'OneCellAnchor'] and nuevo_anchor._from.row >= 12:
                            nuevo_anchor._from.row += filas_extra
                            if tipo_ancla == 'TwoCellAnchor':
                                if hasattr(nuevo_anchor, '_to'):
                                    nuevo_anchor._to.row += filas_extra
                                elif hasattr(nuevo_anchor, 'to'):
                                    nuevo_anchor.to.row += filas_extra
                                    
                        # Clonar imagen en memoria
                        img_bytes = io.BytesIO()
                        imagen.ref.seek(0)
                        img_bytes.write(imagen.ref.read())
                        img_bytes.seek(0)
                        
                        nueva_imagen = OpenpyxlImage(img_bytes)
                        nueva_imagen.anchor = nuevo_anchor
                        nueva_hoja.add_image(nueva_imagen)
                    except Exception as e:
                        logging.warning(f"Aviso al copiar logo o firma: {e}")

                # --- 4. LLENADO DE DATOS Y FORMATO ---
                for index, datos in enumerate(lista_datos):
                    expediente = datos.get("expediente", {})
                    serie = datos.get("info_serie", {}) or {}
                    respuestas = datos.get("respuestas", [])

                    # Cálculo de folios totales
                    total_folios = 0
                    try:
                        if expediente.get("paginas"): total_folios += int(expediente.get("paginas"))
                    except ValueError: pass
                    for r in respuestas:
                        try:
                            if r.get("paginas"): total_folios += int(r.get("paginas"))
                        except ValueError: pass

                    clasificacion = expediente.get("clasificacion", "")
                    num_exp = ""
                    if clasificacion:
                        partes = clasificacion.split('/')
                        if len(partes) >= 3: num_exp = partes[-2]

                    f_actual = fila_inicio + index

                    # 1 a 7: Datos básicos del expediente
                    nueva_hoja.cell(row=f_actual, column=1, value=index + 1)
                    nueva_hoja.cell(row=f_actual, column=2, value=expediente.get("lote_origen", "")) # Caja
                    nueva_hoja.cell(row=f_actual, column=3, value=num_exp)
                    nueva_hoja.cell(row=f_actual, column=4, value=clasificacion)
                    nueva_hoja.cell(row=f_actual, column=5, value=expediente.get("asunto", ""))
                    nueva_hoja.cell(row=f_actual, column=6, value=expediente.get("apertura", ""))
                    nueva_hoja.cell(row=f_actual, column=7, value=expediente.get("cierre", ""))
                    
                    # 8: Número de folios
                    nueva_hoja.cell(row=f_actual, column=8, value=total_folios) 
                    
                    # 9 a 11: Valores Documentales
                    nueva_hoja.cell(row=f_actual, column=9, value=marca(serie.get("administrativo")))
                    nueva_hoja.cell(row=f_actual, column=10, value=marca(serie.get("legal")))
                    nueva_hoja.cell(row=f_actual, column=11, value=marca(serie.get("fiscal")))
                    
                    # 12 a 14: Vigencias
                    nueva_hoja.cell(row=f_actual, column=12, value=serie.get("tramite", ""))
                    nueva_hoja.cell(row=f_actual, column=13, value=serie.get("concentracion", ""))
                    nueva_hoja.cell(row=f_actual, column=14, value=serie.get("total", ""))
                    
                    # 15 a 17: Clasificación de Información
                    nueva_hoja.cell(row=f_actual, column=15, value=marca(serie.get("reservada")))
                    nueva_hoja.cell(row=f_actual, column=16, value=marca(serie.get("confidencial")))
                    nueva_hoja.cell(row=f_actual, column=17, value=marca(serie.get("publica")))
                    
                    # 18 a 21: Ubicación Topográfica
                    nueva_hoja.cell(row=f_actual, column=18, value=expediente.get("ubicacion_area", ""))
                    nueva_hoja.cell(row=f_actual, column=19, value=expediente.get("ubicacion_pasillo", ""))
                    nueva_hoja.cell(row=f_actual, column=20, value=expediente.get("ubicacion_anaquel", ""))
                    nueva_hoja.cell(row=f_actual, column=21, value=expediente.get("ubicacion_charola", ""))
                    
                    # 22 a 24: Tradición y Tomo
                    nueva_hoja.cell(row=f_actual, column=22, value=marca(serie.get("original")))
                    nueva_hoja.cell(row=f_actual, column=23, value=marca(serie.get("copia")))
                    nueva_hoja.cell(row=f_actual, column=24, value=expediente.get("")) # Tomo
                    
                    # Centrado y Bordes de celdas (1 a 24)
                    for col in range(1, 25):
                        celda_actual = nueva_hoja.cell(row=f_actual, column=col)
                        
                        celda_actual.border = borde_fino
                        
                        if col != 5: # Todo centrado menos el Asunto
                            celda_actual.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            celda_actual.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            wb.remove(hoja_plantilla)
            
            wb.save(output_path)
            wb.close()
            
            return True, output_path
            
        except Exception as e:
            logging.error(f"Error al generar inventario primario: {e}", exc_info=True)
            return False, str(e)
        
    def generar_vale_prestamo_desde_plantilla(self, ruta_plantilla: str, clasificacion: str, asunto: str, solicitante: str, area: str, f_prestamo: str, f_vencimiento: str, responsable: str()) -> tuple:
        """
        Abre la plantilla de Excel, busca los marcadores dinámicos, inyecta los datos 
        y guarda una copia lista para imprimir en el Escritorio.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return False, "La librería 'openpyxl' no está instalada. Ejecute: pip install openpyxl"

        if not os.path.exists(ruta_plantilla):
            return False, f"No se encontró la plantilla en: {ruta_plantilla}"

        try:
            wb = load_workbook(ruta_plantilla)
            ws = wb.active

            marcadores = {
                "[CLASIFICACION]": clasificacion,
                "[ASUNTO]": asunto,
                "[SOLICITANTE]": solicitante,
                "[AREA]": area,
                "[FECHA_SALIDA]": f_prestamo,
                "[FECHA_LIMITE]": f_vencimiento,
                "[RESPONSABLE]": responsable
            }

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for tag, valor_real in marcadores.items():
                            if tag in cell.value:
                                # Reemplazamos la etiqueta por el valor de la base de datos
                                cell.value = cell.value.replace(tag, str(valor_real))

            nombre_salida = f"Vale_Prestamo_{clasificacion.replace('/', '-')}.xlsx"
            ruta_salida = os.path.abspath(os.path.join(os.path.expanduser("~"), "Desktop", nombre_salida))
            
            wb.save(ruta_salida)
            return True, ruta_salida

        except Exception as e:
            import logging
            logging.error(f"Error al generar vale Excel: {e}", exc_info=True)
            return False, f"Error al procesar la plantilla de Excel: {e}"